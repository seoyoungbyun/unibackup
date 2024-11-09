#! /usr/bin/python3

from openpyxl import load_workbook

def get_total(r, sheet):
    midterm = sheet.cell(row = r, column = 3).value * 0.3
    final = sheet.cell(row = r, column = 4).value * 0.35
    hw = sheet.cell(row = r, column = 5).value * 0.34
    attendance = sheet.cell(row = r, column = 6).value
    
    return midterm + final + hw + attendance

filepath = 'student.xlsx'
wb = load_workbook(filename = filepath)
sheet = wb.active

totals = []
for r in range(2, sheet.max_row + 1):
    total = get_total(r, sheet)
    totals.append(total)
    sheet.cell(row = r, column = 7, value = total)
totals.sort(reverse=True)

all = len(totals)
print(totals)
a = int(all * 0.3)
aPlus = int(a * 0.5)
b = int(all * 0.7) - a
bPlus = int(b * 0.5)
c = all - a - b
cPlus = int(c * 0.5)


for r in range(2, sheet.max_row + 1):
    grade = ""
    if sheet.cell(row = r, column = 7).value > totals[aPlus]:
        grade = "A+"
    elif sheet.cell(row = r, column = 7).value > totals[a]:
        grade = "A0"
    elif sheet.cell(row = r, column = 7).value > totals[(a + bPlus)]:
        grade = "B+"
    elif sheet.cell(row = r, column = 7).value > totals[(a + b)]:
        grade = "B0"
    elif sheet.cell(row = r, column = 7).value > totals[(a + b + cPlus)]:
        grade = "C+"
    else:
        grade = "C0"

    sheet.cell(row = r, column = 8, value = grade)
wb.save(filepath)